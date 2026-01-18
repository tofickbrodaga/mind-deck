import io
from typing import List, Tuple
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import pytesseract
from fastapi import UploadFile


class ImportService:
    """Сервис для импорта карточек из различных форматов"""
    
    async def import_from_word(self, file: UploadFile) -> List[Tuple[str, str]]:
        """Импортировать карточки из Word документа"""
        content = await file.read()
        doc = Document(io.BytesIO(content))
        
        cards = []
        current_front = None
        
        for paragraph in doc.paragraphs:
            text = paragraph.text.strip()
            if not text:
                continue
            
            # Простая логика: первая строка - термин, вторая - определение
            if current_front is None:
                current_front = text
            else:
                cards.append((current_front, text))
                current_front = None
        
        return cards
    
    async def import_from_excel(self, file: UploadFile) -> List[Tuple[str, str]]:
        """Импортировать карточки из Excel файла"""
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook.active
        
        cards = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2 and row[0] and row[1]:
                front = str(row[0]).strip()
                back = str(row[1]).strip()
                if front and back:
                    cards.append((front, back))
        
        return cards
    
    async def import_from_image(self, file: UploadFile) -> List[Tuple[str, str]]:
        """Импортировать карточки из изображения с помощью OCR"""
        content = await file.read()
        image = Image.open(io.BytesIO(content))
        text = pytesseract.image_to_string(image, lang='rus+eng')

        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        cards = []
        current_front = None
        
        for line in lines:
            if current_front is None:
                current_front = line
            else:
                cards.append((current_front, line))
                current_front = None
        
        return cards
